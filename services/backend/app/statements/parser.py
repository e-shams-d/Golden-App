"""The deterministic statement parser. `08_Bank_File_and_Result_Processing.md` §8.4-8.6.

M10 slice 4. **The parser this repository has been recording the absence of since M2.**
`BANK-VER-005` was filed because "the mappings parse the fixtures" was a claim nothing could
evaluate, and its text was corrected at M8's close when a statement parser had been confused with
an evidence renderer. This is the statement parser. It does not by itself discharge that gap — see
the M10 plan §2 — but it is the piece the gap was waiting on.

**Pure, and that is deliberate.** In: workbook bytes, a mapping, normalization rules. Out: a list
of canonical rows. No session, no storage, no clock. Every rule document 08 §8.5 states is a
property of this function, so every one of them can be tested without a database, and a mapping can
be tried against a real file in a unit test rather than a fixture chain.

**The field names are an allowlist, not a lookup.** `bank_mappings.mapping` is operator-supplied
configuration and `tests/fixtures/bank_fixtures.py` deliberately carries one whose `field` is
`amount_irr"; DROP TABLE bank_mappings; --`. A mapping naming a field this module does not know is
a **configuration error that fails the run** with the name in `error_summary`; it never becomes an
attribute name, a column reference or anything else. Enforcement by absence: the parser can only
produce the fields it declares.

**Nothing is guessed.** §8.4: "Missing fields remain null. They must not be guessed." A cell that
does not parse leaves its normalized field null and its raw string kept, and the row's status says
so. §8.5's "Do not silently convert debit to credit or vice versa" is why deposit and withdrawal
are read from separate mapped columns and never inferred from a sign.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import UTC, datetime
from typing import Any, Final

from openpyxl import load_workbook

from app.core.hashing import unversioned_digest

# The canonical fields a mapping may name. Anything else fails the run.
#
# Spelled as document 04 §10.6 spells the columns rather than as document 08 §8.4 does
# (`deposit_amount_irr`, `withdrawal_amount_irr`): the schema is the authority for column names and
# a parser that produced one vocabulary while the table used another is DOC-CONFLICT-047 repeating
# itself. That conflict cost a rewrite in M2 for exactly this reason.
KNOWN_FIELDS: Final[frozenset[str]] = frozenset(
    {
        "transaction_date",
        "transaction_time",
        "amount_in_irr",
        "amount_out_irr",
        "balance_irr",
        "document_number",
        "tracking_number",
        "description",
        "counterparty_name",
        "counterparty_account",
        "counterparty_iban",
    }
)

# The three the fingerprint and the match index are built from. A mapping that names none of them
# describes a file this platform cannot match anything against, which is a configuration error
# worth refusing before a single row is written rather than after.
REQUIRED_FIELDS: Final[frozenset[str]] = frozenset({"transaction_date", "amount_in_irr"})

ROW_VALID: Final = "valid"
ROW_WARNING: Final = "warning"
ROW_INVALID: Final = "invalid"
ROW_IGNORED_EMPTY: Final = "ignored_empty"

# Persian and Arabic-Indic digits fold to ASCII when the mapping asks. Both ranges, because a
# single bank file routinely contains both — the Persian ۴ and the Arabic ٤ are different code
# points for the same digit.
_PERSIAN_DIGITS = "۰۱۲۳۴۵۶۷۸۹"
_ARABIC_DIGITS = "٠١٢٣٤٥٦٧٨٩"
_DIGIT_FOLD: Final = str.maketrans(
    {ord(char): str(index) for index, char in enumerate(_PERSIAN_DIGITS)}
    | {ord(char): str(index) for index, char in enumerate(_ARABIC_DIGITS)}
)

_JALALI_DATE = re.compile(r"^(\d{4})[/-](\d{1,2})[/-](\d{1,2})$")
_TIME = re.compile(r"^(\d{1,2}):(\d{2})(?::(\d{2}))?$")

# Jalali months: the first six are 31 days, the next five 30, and Esfand is 29 or 30. Used only to
# reject an impossible date, never to convert one — see `_normalized_instant`.
_JALALI_MONTH_LENGTHS: Final = (31, 31, 31, 31, 31, 31, 30, 30, 30, 30, 30, 30)


class MappingConfigurationError(Exception):
    """The mapping cannot parse any file, so no row is worth attempting.

    Distinct from a row that fails to parse: that is data, recorded on the row and visible in
    preview. This is configuration, and document 08 §22.2's answer to it is a new import run after
    the mapping is corrected — not a run that half-succeeds against a mapping nobody fixed.
    """


@dataclass(frozen=True, slots=True)
class ParsedRow:
    """One canonical row. Document 08 §8.4, in the schema's vocabulary."""

    row_number: int
    status: str
    raw_data: dict[str, str]
    row_fingerprint: str
    transaction_at_normalized: datetime | None = None
    transaction_date_raw: str | None = None
    transaction_time_raw: str | None = None
    amount_in_irr: int | None = None
    amount_out_irr: int | None = None
    balance_irr: int | None = None
    document_number: str | None = None
    tracking_number: str | None = None
    description: str | None = None
    counterparty_name: str | None = None
    counterparty_account: str | None = None
    counterparty_iban: str | None = None
    # Why the row is not `valid`. Carried out of the parser so the run's `error_summary` can name
    # the row and the reason, which is §22.2's "preserve import-run errors".
    problems: tuple[str, ...] = ()


@dataclass(frozen=True, slots=True)
class ParseResult:
    rows: tuple[ParsedRow, ...]
    # Columns the bank's file carried that the mapping does not name. Not an error — §22.2 refuses
    # to hide anything — and worth surfacing, because an unmapped column is usually the first sign
    # a bank changed its format.
    unmapped_headers: tuple[str, ...] = ()
    problems: tuple[str, ...] = field(default_factory=tuple)


def parse_statement(
    content: bytes,
    *,
    mapping: dict[str, Any],
    normalization_rules: dict[str, Any] | None = None,
) -> ParseResult:
    """Read the workbook. Raises `MappingConfigurationError` when the mapping cannot apply."""

    columns = _mapped_columns(mapping)
    rules = normalization_rules or {}
    fold_digits = rules.get("digits") == "fold_persian_to_ascii"

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as error:  # openpyxl raises a wide family for a bad file
        raise MappingConfigurationError(
            f"the uploaded file could not be opened as a workbook: {error}"
        ) from error

    try:
        sheet = workbook.worksheets[0]
        grid = [tuple(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    if not grid:
        raise MappingConfigurationError(
            "the workbook has no rows at all, so there is no header to match the mapping against"
        )

    headers = [_text(cell, fold_digits=False) or "" for cell in grid[0]]
    positions = _header_positions(headers, columns)

    missing = sorted(set(columns) - set(positions))
    if missing:
        raise MappingConfigurationError(
            "the statement is missing columns the mapping requires: "
            + ", ".join(f"{columns[header]!r} (header {header!r})" for header in missing)
        )

    unmapped = tuple(
        header for header in headers if header and header not in columns
    )

    rows: list[ParsedRow] = []
    for offset, cells in enumerate(grid[1:], start=1):
        rows.append(
            _parse_row(
                row_number=offset,
                headers=headers,
                cells=cells,
                columns=columns,
                positions=positions,
                fold_digits=fold_digits,
            )
        )

    return ParseResult(rows=tuple(rows), unmapped_headers=unmapped)


def _mapped_columns(mapping: dict[str, Any]) -> dict[str, str]:
    """`{header: field}`, validated against the allowlist.

    Every failure here is a `MappingConfigurationError` rather than a row problem: a mapping that
    names an unknown field is broken for every row in every file, and reporting it per row would
    produce one identical error per line of the statement.
    """

    entries = mapping.get("columns")
    if not isinstance(entries, list) or not entries:
        raise MappingConfigurationError(
            "the mapping has no `columns` list, so it names no part of the bank's file"
        )

    columns: dict[str, str] = {}
    for entry in entries:
        if not isinstance(entry, dict):
            raise MappingConfigurationError(f"mapping column {entry!r} is not an object")
        header = entry.get("header")
        name = entry.get("field")
        if not isinstance(header, str) or not header.strip():
            raise MappingConfigurationError(f"mapping column {entry!r} has no header")
        if not isinstance(name, str) or name not in KNOWN_FIELDS:
            raise MappingConfigurationError(
                f"mapping column {header!r} names field {name!r}, which this parser does not "
                f"produce. Known fields: {', '.join(sorted(KNOWN_FIELDS))}."
            )
        if header in columns:
            raise MappingConfigurationError(
                f"the mapping names header {header!r} twice, so which column feeds "
                f"{name!r} is undecidable"
            )
        columns[header] = name

    named = set(columns.values())
    absent = sorted(REQUIRED_FIELDS - named)
    if absent:
        raise MappingConfigurationError(
            "the mapping names no column for: "
            + ", ".join(absent)
            + ". A statement with neither a date nor an incoming amount cannot be matched "
            "against anything."
        )
    return columns


def _header_positions(headers: list[str], columns: dict[str, str]) -> dict[str, int]:
    positions: dict[str, int] = {}
    for index, header in enumerate(headers):
        if header in columns and header not in positions:
            positions[header] = index
    return positions


def _parse_row(
    *,
    row_number: int,
    headers: list[str],
    cells: tuple[Any, ...],
    columns: dict[str, str],
    positions: dict[str, int],
    fold_digits: bool,
) -> ParsedRow:
    """One source row. Never raises: a row that cannot be read is a row with a status.

    §22.2's fourth requirement — "never partially hide invalid rows" — is why. A parser that
    dropped the rows it could not read would produce a run whose `row_count` silently disagreed
    with the file, and the operator would have no way to see which lines went missing.
    """

    # **Every cell, not only the mapped ones, and unfolded.** §8.5's first rule is "Preserve every
    # raw source value", and folding here would mean the raw copy had already been edited. The
    # folded reading is used for the canonical fields below and nowhere else.
    raw_data = {
        _raw_key(headers, index): text
        for index, cell in enumerate(cells)
        if (text := _text(cell, fold_digits=False)) is not None
    }

    if not raw_data:
        return ParsedRow(
            row_number=row_number,
            status=ROW_IGNORED_EMPTY,
            raw_data={},
            # An empty row still gets a fingerprint, and it includes the row's own position: two
            # blank lines are not duplicates of each other, and a constant would make every blank
            # line in the file collide with every other.
            row_fingerprint=unversioned_digest({"empty_row": row_number}),
        )

    by_field: dict[str, str | None] = {}
    for header, name in columns.items():
        index = positions[header]
        by_field[name] = (
            _text(cells[index], fold_digits=fold_digits) if index < len(cells) else None
        )

    problems: list[str] = []

    amount_in, trouble = _amount(by_field.get("amount_in_irr"))
    if trouble:
        problems.append(f"incoming {trouble}")
    amount_out, trouble = _amount(by_field.get("amount_out_irr"))
    if trouble:
        problems.append(f"outgoing {trouble}")
    balance, trouble = _amount(by_field.get("balance_irr"))
    if trouble:
        problems.append(f"balance {trouble}")

    date_raw = by_field.get("transaction_date")
    time_raw = by_field.get("transaction_time")
    instant, trouble = _normalized_instant(date_raw, time_raw)
    if trouble:
        problems.append(trouble)

    # §8.6's "mutually coherent deposit/withdrawal values", and §8.5's "Do not silently convert
    # debit to credit or vice versa". Both columns carrying a positive number describes a transfer
    # that went two ways at once; the row is flagged rather than being resolved by picking one.
    if amount_in and amount_out:
        problems.append(
            "the row carries both an incoming and an outgoing amount, which describes one "
            "transfer going two ways"
        )

    if amount_in is None and amount_out is None:
        problems.append("the row carries no amount in either direction")

    fingerprint = unversioned_digest(
        {
            # Over the **normalized** values, per §8.4's `normalized_fingerprint`: two rows that
            # describe the same transfer with different spacing or Persian digits must collide,
            # and a fingerprint over `raw_data` would not.
            "amount_in_irr": amount_in,
            "amount_out_irr": amount_out,
            "transaction_at": instant.isoformat() if instant else None,
            "transaction_date_raw": date_raw,
            "tracking_number": by_field.get("tracking_number"),
            "document_number": by_field.get("document_number"),
            "counterparty_iban": by_field.get("counterparty_iban"),
        }
    )

    return ParsedRow(
        row_number=row_number,
        status=_status_for(problems, amount_in=amount_in, amount_out=amount_out),
        raw_data=raw_data,
        row_fingerprint=fingerprint,
        transaction_at_normalized=instant,
        transaction_date_raw=date_raw,
        transaction_time_raw=time_raw,
        amount_in_irr=amount_in,
        amount_out_irr=amount_out,
        balance_irr=balance,
        document_number=by_field.get("document_number"),
        tracking_number=by_field.get("tracking_number"),
        description=by_field.get("description"),
        counterparty_name=by_field.get("counterparty_name"),
        counterparty_account=by_field.get("counterparty_account"),
        counterparty_iban=by_field.get("counterparty_iban"),
        problems=tuple(problems),
    )


def _raw_key(headers: list[str], index: int) -> str:
    """The header a raw value is filed under, or its position when there is none.

    A column with a blank header still holds data, and dropping it would be the hiding §22.2
    refuses. `column_7` is a position, unambiguous, and visibly not something a bank wrote.
    """

    if index < len(headers) and headers[index]:
        return headers[index]
    return f"column_{index + 1}"


def _status_for(problems: list[str], *, amount_in: int | None, amount_out: int | None) -> str:
    """§8.6's states, and the line between two of them.

    `invalid` when the row cannot be used for matching at all — no amount, or an amount that is
    not a number. `warning` when it can be, but something about it needs a human: a Jalali date
    left unconverted is the common case, and such a row is still matchable by amount and tracking
    number.

    A row is never silently `valid` when it carries a problem, which is the whole difference
    between a preview an accountant can trust and one that reads as clean because nothing looked.
    """

    if not problems:
        return ROW_VALID
    if amount_in is None and amount_out is None:
        return ROW_INVALID
    return ROW_WARNING


def _text(value: Any, *, fold_digits: bool) -> str | None:
    """A cell as a string, or None when it holds nothing.

    Datetimes are rendered ISO rather than `str()`, so a cell openpyxl typed as a date and one the
    bank wrote as text produce the same raw string.
    """

    if value is None:
        return None
    if isinstance(value, datetime):
        text = value.isoformat()
    elif isinstance(value, bool):
        text = "true" if value else "false"
    elif isinstance(value, float) and value.is_integer():
        # openpyxl types every unformatted number as a float. An integral one is written without
        # the `.0` so an amount reads as an amount.
        text = str(int(value))
    else:
        text = str(value)
    text = text.strip()
    if not text:
        return None
    return text.translate(_DIGIT_FOLD) if fold_digits else text


def _amount(text: str | None) -> tuple[int | None, str | None]:
    """An IRR amount as an integer, or a reason it is not one.

    §8.5: "Reject or flag decimal/fractional IRR unless explicitly supported." Rejected here — IRR
    has no subunit in this system, every other amount column is `BigInteger`, and a bank file
    carrying `1000.5` means something nobody has decided.
    """

    if text is None:
        return None, None
    cleaned = text.replace(",", "").replace("٬", "").replace("،", "").strip()
    if cleaned.startswith("+"):
        cleaned = cleaned[1:]
    if not cleaned:
        return None, None
    if cleaned.endswith(".0"):
        cleaned = cleaned[:-2]
    try:
        return int(cleaned), None
    except ValueError:
        return None, f"amount {text!r} is not a whole number of rial"


def _normalized_instant(date_text: str | None, time_text: str | None) -> tuple[
    datetime | None, str | None
]:
    """A timezone-aware instant, or a reason there is none.

    **Jalali dates are validated and not converted.** ADR-006 makes the business timezone
    configuration and the calendar conversion a decision this milestone has no mandate to make; a
    parser that guessed a Gregorian equivalent would write an instant nobody approved into the
    column slice 5 matches on. So a Jalali date is recognised, checked for plausibility, kept
    verbatim in `transaction_date_raw`, and leaves the instant null — which §8.4 explicitly
    permits and this row's status records.

    A Gregorian ISO date is converted, because that needs no calendar decision.
    """

    if date_text is None:
        return None, "the row has no transaction date"

    hour, minute, second = 0, 0, 0
    if time_text is not None:
        match = _TIME.match(time_text)
        if match is None:
            return None, f"time {time_text!r} is not readable"
        hour, minute = int(match.group(1)), int(match.group(2))
        second = int(match.group(3) or 0)
        if hour > 23 or minute > 59 or second > 59:
            return None, f"time {time_text!r} is not a time of day"

    try:
        parsed = datetime.fromisoformat(date_text)
    except ValueError:
        parsed = None

    if parsed is not None:
        instant = parsed if parsed.tzinfo else parsed.replace(tzinfo=UTC)
        if time_text is not None:
            instant = instant.replace(hour=hour, minute=minute, second=second, microsecond=0)
        return instant, None

    jalali = _JALALI_DATE.match(date_text)
    if jalali is None:
        return None, f"date {date_text!r} is neither an ISO date nor a Jalali one"

    year, month, day = (int(group) for group in jalali.groups())
    if not 1 <= month <= 12 or day < 1 or day > _JALALI_MONTH_LENGTHS[month - 1]:
        return None, f"date {date_text!r} is not a date in the Jalali calendar"
    if year < 1300 or year > 1500:
        return None, f"date {date_text!r} has an implausible Jalali year"

    return None, (
        "the date is Jalali and is preserved raw; ADR-006 leaves the calendar conversion "
        "undecided, so no instant is invented"
    )
