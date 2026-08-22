"""The bank file itself. `15_Agent_Implementation_Plan.md` §15.6.

M7 slice 2. This module takes **data and returns bytes**. No session, no ORM row, no database —
the same shape `app/batching/splitting.py` takes its rules in, and for the same reason: the rules
about what a bank file must contain are worth testing without a database, and a renderer that
took a `Session` could only be exercised by a test that had one.

**Four properties §15.6 requires, and one it does not state in those words.**

- *Amounts are integers.* `:975` — "never use floating-point values for amounts". Written as
  Python `int` and asserted by reading the cell's type back out of the file, because a float that
  happens to print as `1500000000` is indistinguishable from an integer until it is not.
- *Row order is the version's.* `:972-974`. The caller passes rows already ordered and this
  writes them in that order; it does not sort, because a renderer that sorted would be deciding
  something the version already decided.
- *Persian and English both survive.* xlsx is XML in a zip and the encoding is UTF-8 throughout,
  so this costs nothing — but it is asserted rather than assumed, because "the names came out as
  question marks" is discovered by a bank, not by us.
- *Formula-like text is inert.* `:971` — "escape or reject formula-like untrusted text according
  to policy". See below; this is G-6 and the policy is stated here rather than implied.
- And the one §15.6 does not say: **the untrusted text is the beneficiary name and the
  description, both typed by a trader.** Everything else in the row is a number this system
  generated or an IBAN a CHECK constrains to `^IR[0-9]{24}$`. Knowing which cells are dangerous is
  what makes the escaping narrow enough to reason about.

**G-6, decided by measurement rather than by preference.** The question was escape or refuse, and
the plan framed the cost of escaping as "changes what a payee name looks like". That framing came
from the usual mitigation — prefixing an apostrophe — and openpyxl stores that apostrophe *inside
the value*, so a bank would read `'=HYPERLINK(…)` including the quote. Measured, three options
behave differently:

| written as | stored as a formula | payee name altered |
|---|---|---|
| the raw string | **yes**, `<f>` in the sheet XML | no |
| `"'" + value` | no | **yes** |
| explicit string type **and** `quotePrefix` | no | no |

The third is what this module does. `quotePrefix` is Excel's own marker for "this text is text" —
the thing the apostrophe imitates — so the cell displays and exports exactly what the trader typed
while never being evaluated, and it stays text when somebody opens the file and edits around it.
**Nothing is refused and nothing is altered**, which makes G-6's trade-off mostly disappear rather
than resolving it in one direction. The owner may still prefer refusal; `SEC-EXPORT-001` asserts
the written cell, so changing the policy is a change here and a change to that assertion.

**A formula is not the only dangerous prefix.** `=` is the obvious one; `+`, `-` and `@` are all
treated as formula introducers by Excel, and `\t`, `\r` and `\n` can smuggle one past a naive
check that looks only at the first character. The guard strips nothing — stripping would alter the
name — it decides *by the first non-whitespace character* whether the cell needs the marker.
"""

from __future__ import annotations

import io
from collections.abc import Sequence
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# Excel evaluates a cell whose text begins with any of these. `@` is the legacy `DDE`/function
# introducer and is included for the same reason the other three are.
FORMULA_INTRODUCERS: frozenset[str] = frozenset({"=", "+", "-", "@"})

# Characters Excel ignores before deciding a cell is a formula, so a value beginning
# "\t=cmd(...)" is a formula despite not starting with one of the four above. Checked rather than
# stripped: the beneficiary's name is evidence and this module does not edit evidence.
LEADING_NOISE = " \t\r\n\v\f ‎‏"


@dataclass(frozen=True, slots=True)
class ExportRow:
    """One payment, already ordered and already snapshotted.

    Every field is a value the batch version froze — `04_Database_Schema.md:1021-1023` calls the
    beneficiary fields "the exact approved/exported value". Nothing here is looked up, because a
    lookup at render time could return something the manager never approved.
    """

    row_order: int
    beneficiary_name: str
    beneficiary_iban: str
    amount_irr: int
    description: str | None = None

    def __post_init__(self) -> None:
        if not isinstance(self.amount_irr, int) or isinstance(self.amount_irr, bool):
            # `bool` is an `int` in Python and would write `TRUE` into an amount column. The
            # check is cheap and the failure it prevents is a bank file nobody can process.
            raise TypeError(
                f"row {self.row_order}: amount_irr must be int, not "
                f"{type(self.amount_irr).__name__}. 15_Agent_Implementation_Plan.md:975 forbids "
                "floating-point amounts, and a float that prints like an integer is "
                "indistinguishable from one until a bank rejects the file."
            )
        if self.amount_irr <= 0:
            raise ValueError(f"row {self.row_order}: amount_irr must be positive")


# The columns, in order, with the header each carries. Fixed here rather than driven by
# `bank_mappings.mapping`, which is empty in every fixture this milestone has and whose schema no
# document pins down. Slice 3 owns the mapping-driven form if the owner's real bank templates
# need one (ADR-007); until then a fixed set is honest and a mapping-shaped one that ignores the
# mapping would not be.
HEADERS: tuple[str, ...] = ("ردیف", "نام ذی‌نفع", "شبا", "مبلغ (ریال)", "شرح")


def needs_text_marker(value: str) -> bool:
    """Whether Excel would evaluate this cell's text as a formula."""

    stripped = value.lstrip(LEADING_NOISE)
    return bool(stripped) and stripped[0] in FORMULA_INTRODUCERS


def _write_text(sheet: Worksheet, row: int, column: int, value: str | None) -> None:
    """Write untrusted text so it is displayed exactly and evaluated never.

    Two things together, and each alone is insufficient. `data_type = "s"` stops openpyxl
    inferring a formula from the leading character; `quotePrefix` is the cell format that tells
    Excel the same thing, so the file survives being opened and edited. The value is untouched.
    """

    cell = sheet.cell(row=row, column=column)
    if value is None:
        return
    cell.value = value
    if needs_text_marker(value):
        cell.data_type = "s"
        cell.quotePrefix = True


def render_bank_file(rows: Sequence[ExportRow], *, sheet_title: str = "پرداخت‌ها") -> bytes:
    """One sheet, one header row, then the payments in the order given.

    Returns bytes rather than writing a path, so the caller decides where it lands and — more to
    the point — can hash what it is about to store rather than hashing a file it hopes is the
    same one. `FINANCIAL_INTEGRITY_BASELINE.md` §1 requires the record to be written only after
    the artifact verifies, and that is easier to get right when the artifact is a value.
    """

    if not rows:
        # An empty bank file is a file that instructs a bank to do nothing, and there is no
        # legitimate way to reach one: `ck_payment_batch_versions_row_count` already refuses a
        # version with no rows. Raising here rather than writing it makes the two agree.
        raise ValueError("a bank file with no rows cannot be rendered")

    book = Workbook()
    sheet = book.active
    if sheet is None:  # pragma: no cover - a new Workbook always has one
        raise RuntimeError("openpyxl returned a workbook with no active sheet")
    sheet.title = sheet_title

    for column, header in enumerate(HEADERS, start=1):
        sheet.cell(row=1, column=column, value=header)

    for offset, row in enumerate(rows, start=2):
        sheet.cell(row=offset, column=1, value=row.row_order)
        _write_text(sheet, offset, 2, row.beneficiary_name)
        _write_text(sheet, offset, 3, row.beneficiary_iban)
        # Not through `_write_text`: an amount is an integer and must stay one. Passing it
        # through the text path would write it as a string and `SVC-EXPORT-003` would fail,
        # which is the correct outcome — the two paths are separate because the types are.
        sheet.cell(row=offset, column=4, value=row.amount_irr)
        _write_text(sheet, offset, 5, row.description)

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()
