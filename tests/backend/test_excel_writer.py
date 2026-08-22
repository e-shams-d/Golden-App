"""The bank file, asserted by reading the file back — never by trusting what was passed in.

M7 slice 2. `15_Agent_Implementation_Plan.md` §15.6 makes four demands of this file, and every one
of them is the kind a test written against the *input* would pass while the output was wrong. So
every assertion here opens the produced bytes: with openpyxl for values and types, and `zipfile`
for the sheet XML where the distinction between a string and a formula actually lives.

**Why the XML and not just the value.** `load_workbook` returns `'=HYPERLINK(…)'` for a cell
stored as a formula *and* for one stored as text — the string is the same either way. Only
`<f>` in `xl/worksheets/sheet1.xml` says which it is, and that is the difference between a bank
opening a spreadsheet and a bank opening a spreadsheet that runs something.

Covers: SVC-EXPORT-003, SVC-EXPORT-004, SEC-EXPORT-001.
"""

from __future__ import annotations

import io
import zipfile

import pytest
from app.exports.excel import ExportRow, needs_text_marker, render_bank_file
from openpyxl import load_workbook

PERSIAN_NAME = "علی رضایی"
ENGLISH_NAME = "Ali Rezaei"
IBAN = "IR060120000000000000000044"

HYPERLINK = '=HYPERLINK("http://evil.example","click")'


def a_row(**overrides: object) -> ExportRow:
    fields: dict[str, object] = {
        "row_order": 1,
        "beneficiary_name": PERSIAN_NAME,
        "beneficiary_iban": IBAN,
        "amount_irr": 1_500_000_000,
        "description": "بابت خرید",
    }
    fields.update(overrides)
    return ExportRow(**fields)  # type: ignore[arg-type]


def sheet_of(data: bytes) -> object:
    book = load_workbook(io.BytesIO(data))
    return book[book.sheetnames[0]]


def sheet_xml(data: bytes) -> str:
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        return archive.read("xl/worksheets/sheet1.xml").decode("utf-8")


def test_amounts_are_written_as_integers_and_read_back_as_integers() -> None:
    """`SVC-EXPORT-003`. `15_Agent_Implementation_Plan.md:975` forbids floating-point amounts.

    The type is asserted, not the printed value. `1500000000.0` and `1500000000` render
    identically in most viewers, and a float only reveals itself at a scale where it has already
    lost a rial — which for IRR amounts is not a hypothetical scale.
    """

    data = render_bank_file([a_row(amount_irr=1_500_000_000)])
    cell = sheet_of(data).cell(row=2, column=4)  # type: ignore[attr-defined]

    assert isinstance(cell.value, int)
    assert not isinstance(cell.value, float)
    assert cell.value == 1_500_000_000


def test_a_large_amount_keeps_every_rial() -> None:
    """The scale that makes the type matter rather than being a formality.

    12,345,678,901,234 rial is about 12 trillion — a number this platform can reach, and one a
    `float64` cannot hold exactly once it grows a little further. If amounts ever become floats,
    this is the test that says so before a bank does.
    """

    data = render_bank_file([a_row(amount_irr=12_345_678_901_234)])

    assert sheet_of(data).cell(row=2, column=4).value == 12_345_678_901_234  # type: ignore[attr-defined]


def test_a_float_amount_is_refused_at_construction() -> None:
    """Before the file, not after. A renderer that quietly coerced would be worse than one that
    refused: the file would look right and the amount would be wrong by a fraction nobody reads.
    """

    with pytest.raises(TypeError, match="must be int"):
        ExportRow(
            row_order=1,
            beneficiary_name=ENGLISH_NAME,
            beneficiary_iban=IBAN,
            amount_irr=1_500_000_000.0,  # type: ignore[arg-type]
        )


def test_a_boolean_amount_is_refused_too() -> None:
    """`bool` is a subclass of `int` in Python, so a plain `isinstance` check admits `True`.

    It would be written as `TRUE` in the amount column. Worth its own test because the guard that
    catches it is one clause long and reads like defensive noise until you know why it is there.
    """

    with pytest.raises(TypeError, match="must be int"):
        ExportRow(
            row_order=1,
            beneficiary_name=ENGLISH_NAME,
            beneficiary_iban=IBAN,
            amount_irr=True,  # type: ignore[arg-type]
        )


def test_persian_and_english_both_survive_the_round_trip() -> None:
    """`SVC-EXPORT-004`. `15_Agent_Implementation_Plan.md:972-974`.

    Asserted because "the names came out as question marks" is a defect discovered by a bank
    rather than by us, and because the sheet title is Persian too — a title is stored in a
    different part of the archive from the cells and can fail on its own.
    """

    data = render_bank_file(
        [
            a_row(row_order=1, beneficiary_name=PERSIAN_NAME),
            a_row(row_order=2, beneficiary_name=ENGLISH_NAME),
        ]
    )
    sheet = sheet_of(data)

    assert sheet.title == "پرداخت‌ها"  # type: ignore[attr-defined]
    assert sheet.cell(row=2, column=2).value == PERSIAN_NAME  # type: ignore[attr-defined]
    assert sheet.cell(row=3, column=2).value == ENGLISH_NAME  # type: ignore[attr-defined]
    assert sheet.cell(row=2, column=5).value == "بابت خرید"  # type: ignore[attr-defined]


def test_rows_are_written_in_the_order_given_and_never_sorted() -> None:
    """`SVC-EXPORT-004`'s other half. The version decided the order; this does not revisit it.

    The rows are passed deliberately out of ascending `row_order` so that a renderer which
    "helpfully" sorted would produce a different file from the one asserted here. What must
    match the bank file is the version's own ordering, and a renderer is not where that is
    decided.
    """

    data = render_bank_file(
        [
            a_row(row_order=3, beneficiary_name="سوم"),
            a_row(row_order=1, beneficiary_name="اول"),
            a_row(row_order=2, beneficiary_name="دوم"),
        ]
    )
    sheet = sheet_of(data)

    assert [sheet.cell(row=r, column=1).value for r in (2, 3, 4)] == [3, 1, 2]  # type: ignore[attr-defined]
    assert [sheet.cell(row=r, column=2).value for r in (2, 3, 4)] == [  # type: ignore[attr-defined]
        "سوم",
        "اول",
        "دوم",
    ]


def test_a_formula_like_name_is_stored_as_text_and_not_as_a_formula() -> None:
    """`SEC-EXPORT-001`. The assertion is on the written file's cell, not on the value passed in.

    `15_Agent_Implementation_Plan.md:971` says "escape or reject formula-like untrusted text
    according to policy", and the policy this slice implements is in `app/exports/excel.py`:
    mark the cell as text and let the value through unaltered. So two things are asserted
    together, and either alone would be a weaker claim than it looks:

    - the sheet XML contains **no** `<f>` element, which is the only place the difference between
      a string and a formula is recorded;
    - the value the bank reads is **exactly** what the trader typed, because the point of not
      refusing the row is that the payee keeps their name.
    """

    data = render_bank_file([a_row(beneficiary_name=HYPERLINK)])

    assert "<f>" not in sheet_xml(data), "the beneficiary name was written as a formula"
    assert sheet_of(data).cell(row=2, column=2).value == HYPERLINK  # type: ignore[attr-defined]


@pytest.mark.parametrize(
    "dangerous",
    [
        '=HYPERLINK("http://evil.example","click")',
        "+1234567890",
        "-2+3",
        "@SUM(A1:A9)",
        # Leading whitespace does not save a naive check: Excel skips it before deciding.
        "\t=cmd|'/c calc'!A1",
        "  =1+1",
    ],
)
def test_every_formula_introducer_is_neutralised(dangerous: str) -> None:
    """Four introducers and two ways to hide one, each asserted against the written file.

    Parametrised rather than combined into one assertion, for the reason M6's hash test learned:
    a single test that accepted any of several conditions passed with either guard removed.
    """

    data = render_bank_file([a_row(beneficiary_name=dangerous)])

    assert "<f>" not in sheet_xml(data), f"{dangerous!r} was written as a formula"
    assert sheet_of(data).cell(row=2, column=2).value == dangerous  # type: ignore[attr-defined]


def test_an_ordinary_name_is_not_marked() -> None:
    """The other direction, and it is what stops the guard being trivially satisfiable.

    A renderer that marked *every* cell would pass every assertion above while adding a quote
    prefix to every payee in the file. `needs_text_marker` is asserted directly here because the
    absence of a marker is not visible in the value.
    """

    assert not needs_text_marker(PERSIAN_NAME)
    assert not needs_text_marker(ENGLISH_NAME)
    assert not needs_text_marker(IBAN)
    assert not needs_text_marker("")
    assert needs_text_marker(HYPERLINK)


def test_a_description_is_escaped_as_well_as_a_name() -> None:
    """Both untrusted columns, because both are typed by a trader.

    The beneficiary name is the obvious one and the description is the one a guard written from
    the obvious case forgets. `05_API_Specification.md`'s request body carries both from the
    trader unchanged.
    """

    data = render_bank_file([a_row(description="=1+1")])

    assert "<f>" not in sheet_xml(data)
    assert sheet_of(data).cell(row=2, column=5).value == "=1+1"  # type: ignore[attr-defined]


def test_an_empty_file_is_refused() -> None:
    """A bank file with no rows instructs a bank to do nothing, and nothing legitimately produces
    one: `ck_payment_batch_versions_row_count` already refuses a version with no rows. Refusing
    here makes the renderer agree with the database rather than quietly disagreeing.
    """

    with pytest.raises(ValueError, match="no rows"):
        render_bank_file([])


def test_the_header_row_is_present_and_first() -> None:
    """Five columns, and the payments start at row two.

    Asserted because every other test in this file indexes from row 2, and if the header were
    dropped they would all silently start reading the wrong row and still pass.
    """

    data = render_bank_file([a_row()])
    sheet = sheet_of(data)

    assert [sheet.cell(row=1, column=c).value for c in range(1, 6)] == [  # type: ignore[attr-defined]
        "ردیف",
        "نام ذی‌نفع",
        "شبا",
        "مبلغ (ریال)",
        "شرح",
    ]
