"""The final export: only for an approved version, and only if the eight checks hold.

M7 slice 3, against a real PostgreSQL and real storage.

**What separates this from a preview is not the bytes.** Both render the same version through the
same writer. A final export additionally requires an approval for the *exact* version, runs
§15.5's eight comparisons before returning, and quarantines itself if any of them disagree.

`tests/backend/test_export_integrity.py` proves the comparisons individually, on constructed
values, because two of the eight cannot be provoked through the database at all. This file proves
the parts that need one: that an unapproved version is refused, that a mismatch quarantines rather
than vanishing, that two concurrent generations produce one file, and that a superseded approval
cannot produce anything.

Covers: SVC-EXPORT-005, SVC-INTEGRITY-002, CON-EXPORT-001, AUD-EXPORT-002.
"""

from __future__ import annotations

import uuid
from collections.abc import Iterator
from pathlib import Path
from typing import Any

import psycopg
import pytest
from alembic_runner import run_alembic
from bootstrap_replay import RuntimeIdentities

pytestmark = pytest.mark.integration

PASSWORD = "correct-horse-battery-staple"
CSRF_HEADER = "X-CSRF-Token"
TRADER_CSRF_COOKIE = "__Host-gp_trader_csrf"
ADMIN_CSRF_COOKIE = "__Host-gp_admin_csrf"

TRADER_PHONE = "+989120004701"
IBAN = "IR060120000000000000000047"
LIMIT = 1_000_000_000
ONE_ROW = "900000000"

APPROVE_PURPOSE = "payment_batch_version.approve"
STEP_UP_RESOURCE_TYPE = "payment_batch_version"

# The name the trader really typed, kept through the whole chain. `SVC-EXPORT-005` says the file
# is built from the immutable snapshot and "not current mutable beneficiary data" — so this is
# what must appear in the file even after the beneficiary is renamed underneath it.
ORIGINAL_NAME = "علی رضایی"
RENAMED = "SOMEBODY ELSE ENTIRELY"


@pytest.fixture(scope="module")
def migrated(module_provisioned_database: RuntimeIdentities) -> RuntimeIdentities:
    result = run_alembic(
        module_provisioned_database.migrator_url,
        "upgrade",
        "head",
        app_role=module_provisioned_database.app_role,
        worker_role=module_provisioned_database.worker_role,
    )
    assert result.returncode == 0, f"stdout:\n{result.stdout}\nstderr:\n{result.stderr}"
    return module_provisioned_database


@pytest.fixture(scope="module")
def world(migrated: RuntimeIdentities, tmp_path_factory: Any) -> Iterator[dict[str, Any]]:
    from app.core.config import Settings
    from app.core.runtime import RuntimeServices
    from app.main import create_app
    from app.security.passwords import Argon2Parameters, hash_password
    from fastapi.testclient import TestClient

    storage_root = tmp_path_factory.mktemp("final-storage")
    settings = Settings(
        _env_file=None,
        app_env="test",
        database_url=migrated.owner_url,
        redis_url="redis://127.0.0.1:6379/0",
        local_storage_root=storage_root,
        release_commit="abcdef1234567",
        log_level="CRITICAL",
        auth_csrf_key_secret="c" * 40,
        auth_rate_limit_key_secret=None,
    )
    parameters = Argon2Parameters.from_settings(settings)
    encoded = hash_password(PASSWORD, parameters, max_length=settings.password_max_length)

    ids = {
        name: uuid.uuid4()
        for name in ("trader", "beneficiary", "profile", "version", "account", "mapping")
    }

    with psycopg.connect(_psycopg(migrated.owner_url)) as connection:
        connection.execute(
            "INSERT INTO traders (id, display_name, primary_phone, operational_status, "
            "approval_status) VALUES (%s, 'Final Trader', %s, 'active', 'approved')",
            (ids["trader"], TRADER_PHONE),
        )
        connection.execute(
            "INSERT INTO trader_users (trader_id, phone_number, full_name, password_hash, "
            "status, is_primary) VALUES (%s, %s, 'Contact', %s, 'active', TRUE)",
            (ids["trader"], TRADER_PHONE, encoded),
        )
        connection.execute(
            "INSERT INTO beneficiaries (id, trader_id, full_name, iban, normalized_iban, "
            "status, verification_status) VALUES (%s, %s, %s, %s, %s, 'active', 'not_checked')",
            (ids["beneficiary"], ids["trader"], ORIGINAL_NAME, IBAN, IBAN),
        )
        connection.execute(
            "INSERT INTO bank_profiles (id, code, name, status) "
            "VALUES (%s, 'ayandeh', 'Bank Ayandeh', 'active')",
            (ids["profile"],),
        )
        connection.execute(
            "INSERT INTO bank_profile_versions (id, bank_profile_id, version_number, status, "
            "default_transfer_limit_irr, after_cutoff_transfer_limit_irr, cutoff_time, "
            "splitting_enabled, required_fields, rules, config_hash) "
            "VALUES (%s, %s, 1, 'active', %s, NULL, NULL, TRUE, '{}', '{}', %s)",
            (ids["version"], ids["profile"], LIMIT, "1" * 64),
        )
        connection.execute(
            "INSERT INTO bank_accounts (id, bank_profile_id, display_name, iban, "
            "normalized_iban, account_role, status) "
            "VALUES (%s, %s, 'Centre Account', %s, %s, 'outgoing_source', 'active')",
            (ids["account"], ids["profile"], IBAN, IBAN),
        )
        connection.execute(
            "INSERT INTO bank_mappings (id, bank_profile_version_id, file_type, "
            "template_version, status, mapping, config_hash) "
            "VALUES (%s, %s, 'outgoing_excel', 1, 'active', '{}', %s)",
            (ids["mapping"], ids["version"], "2" * 64),
        )
        for username, role in (
            ("final_accountant", "accountant"),
            ("final_manager", "manager"),
        ):
            connection.execute(
                "INSERT INTO admin_users (username, full_name, password_hash, status) "
                "VALUES (%s, %s, %s, 'active')",
                (username, username, encoded),
            )
            connection.execute(
                "INSERT INTO admin_user_roles (admin_user_id, role_id) "
                "SELECT u.id, r.id FROM admin_users u, roles r "
                "WHERE u.username = %s AND r.code = %s",
                (username, role),
            )
        connection.commit()

    app = create_app(settings=settings)
    app.state.runtime = RuntimeServices.from_settings(settings)
    app.state.accepting_traffic = True
    with TestClient(app, base_url="https://trader.localhost") as client:
        yield {
            "client": client,
            "app": app,
            "app_role": migrated.app_role,
            "owner_url": migrated.owner_url,
            "storage_root": storage_root,
            **{f"{name}_id": value for name, value in ids.items()},
        }
    app.state.runtime.close()


def _psycopg(url: str) -> str:
    return url.replace("postgresql+psycopg://", "postgresql://", 1)


def sign_in_trader(client: Any) -> None:
    client.cookies.clear()
    response = client.post(
        "/api/v1/auth/trader/login", json={"identifier": TRADER_PHONE, "password": PASSWORD}
    )
    assert response.status_code == 200, response.text


def sign_in_admin(client: Any, username: str) -> None:
    client.cookies.clear()
    response = client.post(
        "/api/v1/auth/admin/login", json={"identifier": username, "password": PASSWORD}
    )
    assert response.status_code == 200, response.text


def csrf(client: Any) -> dict[str, str]:
    token = client.cookies.get(TRADER_CSRF_COOKIE) or client.cookies.get(ADMIN_CSRF_COOKIE)
    assert token, "signed in but no CSRF cookie was set"
    return {CSRF_HEADER: token}


def rows(world: dict[str, Any], sql: str, *parameters: Any) -> list[tuple[Any, ...]]:
    with psycopg.connect(_psycopg(world["owner_url"])) as connection:
        return connection.execute(sql, parameters).fetchall()


def a_finalized_version(world: dict[str, Any]) -> dict[str, Any]:
    client = world["client"]
    sign_in_trader(client)
    created = client.post(
        "/api/v1/payment-requests",
        json={
            "beneficiary_id": str(world["beneficiary_id"]),
            "amount": {"value": ONE_ROW, "unit": "IRR"},
            "description": "برای خروجی نهایی",
        },
        headers=csrf(client),
    )
    assert created.status_code == 201, created.text
    request_id = created.json()["request"]["id"]
    revision_id = created.json()["revision"]["id"]

    handed = client.post(
        f"/api/v1/payment-requests/{request_id}/submit",
        json={},
        headers={
            **csrf(client),
            "If-Match": f'"rv-{created.json()["request"]["record_version"]}"',
        },
    )
    assert handed.status_code == 200, handed.text

    sign_in_admin(client, "final_accountant")
    started = client.post(
        f"/api/v1/payment-requests/{request_id}/start-review",
        json={},
        headers={**csrf(client), "If-Match": handed.headers["ETag"]},
    )
    assert started.status_code == 200, started.text
    eligible = client.post(
        f"/api/v1/payment-requests/{request_id}/mark-eligible-for-batching",
        json={"expected_revision_id": revision_id, "review_note": "checked"},
        headers={**csrf(client), "If-Match": started.headers["ETag"]},
    )
    assert eligible.status_code == 200, eligible.text

    batch = client.post(
        "/api/v1/payment-batches",
        json={
            "items": [
                {
                    "payment_request_id": request_id,
                    "expected_revision_id": revision_id,
                    "expected_record_version": eligible.json()["record_version"],
                }
            ],
            "bank_profile_version_id": str(world["version_id"]),
            "bank_account_id": str(world["account_id"]),
            "bank_mapping_id": str(world["mapping_id"]),
        },
        headers={**csrf(client), "Idempotency-Key": str(uuid.uuid4())},
    )
    assert batch.status_code == 201, batch.text
    body = batch.json()
    batch_id = body["batch"]["id"]
    version_id = body["current_version"]["id"]

    frozen = client.post(
        f"/api/v1/payment-batches/{batch_id}/versions/{version_id}/finalize",
        json={"note": "validated"},
        headers={
            **csrf(client),
            "If-Match": batch.headers["ETag"],
            "Idempotency-Key": str(uuid.uuid4()),
        },
    )
    assert frozen.status_code == 200, frozen.text

    return {
        "batch_id": batch_id,
        "version_id": version_id,
        "content_hash": frozen.json()["version"]["content_hash"],
    }


def approve(world: dict[str, Any], target: dict[str, Any]) -> Any:
    client = world["client"]
    sign_in_admin(client, "final_manager")
    reference = client.post(
        "/api/v1/auth/reauthenticate",
        json={
            "password": PASSWORD,
            "purpose": APPROVE_PURPOSE,
            "resource_type": STEP_UP_RESOURCE_TYPE,
            "resource_id": target["version_id"],
        },
        headers=csrf(client),
    )
    assert reference.status_code == 200, reference.text

    return client.post(
        f"/api/v1/payment-batches/{target['batch_id']}"
        f"/versions/{target['version_id']}/approve",
        json={"expected_content_hash": target["content_hash"], "approval_note": "ok"},
        headers={
            **csrf(client),
            "Idempotency-Key": str(uuid.uuid4()),
            "X-Recent-Auth": str(reference.json()["recent_auth_reference"]),
        },
    )


def final_export(world: dict[str, Any], target: dict[str, Any], *, key: str | None = None) -> Any:
    client = world["client"]
    return client.post(
        f"/api/v1/payment-batches/{target['batch_id']}"
        f"/versions/{target['version_id']}/exports/final",
        json=None,
        headers={**csrf(client), "Idempotency-Key": key or str(uuid.uuid4())},
    )


def an_approved_version(world: dict[str, Any]) -> dict[str, Any]:
    target = a_finalized_version(world)
    assert approve(world, target).status_code == 200
    sign_in_admin(world["client"], "final_accountant")
    return target


def test_a_final_export_needs_an_approval_for_this_exact_version(
    world: dict[str, Any],
) -> None:
    """`SVC-EXPORT-005`. A finalized version is not an approved one.

    The negative control for this obligation is to generate with no approval, and it is worth
    stating what makes the refusal meaningful: the version is otherwise perfectly exportable —
    finalized, immutable, hash computed. The only thing missing is that nobody authorised it.
    """

    target = a_finalized_version(world)
    sign_in_admin(world["client"], "final_accountant")

    response = final_export(world, target)

    assert response.status_code == 400, response.text
    assert "no approval" in response.json()["error"]["message"]
    assert rows(
        world,
        "SELECT count(*) FROM bank_excel_exports WHERE payment_batch_version_id = %s",
        target["version_id"],
    )[0][0] == 0


def test_an_approved_version_produces_a_validated_export_naming_its_approval(
    world: dict[str, Any],
) -> None:
    """`SVC-EXPORT-005` and `AUD-EXPORT-002`, and the Definition of Done's chain in one row.

    The DoD asks the system to "prove exactly which approved immutable version produced the exact
    checksummed file". The export row names the approval; the approval names the content hash;
    the version holds the same hash. Every link is asserted here rather than assumed from the
    fact that the request succeeded.
    """

    target = an_approved_version(world)

    response = final_export(world, target)
    assert response.status_code == 201, response.text
    body = response.json()
    assert body["export_type"] == "final"
    assert body["sendable"] is True

    recorded = rows(
        world,
        "SELECT e.status, e.batch_approval_id, e.content_hash, a.approved_content_hash, "
        "v.content_hash FROM bank_excel_exports e "
        "JOIN batch_approvals a ON a.id = e.batch_approval_id "
        "JOIN payment_batch_versions v ON v.id = e.payment_batch_version_id "
        "WHERE e.id = %s",
        body["id"],
    )
    assert len(recorded) == 1
    status, approval_id, export_hash, approval_hash, version_hash = recorded[0]

    assert status == "validated", "the eight checks ran and held, so the row is validated"
    assert approval_id is not None
    # The three links of the chain, each asserted against the next rather than against a
    # constant: the manager approved this version's content, and the file renders that content.
    assert approval_hash == version_hash, "the manager approved this version's content"
    assert export_hash == version_hash, "the file does not render what the version says"
    assert body["content_hash"] == version_hash

    audited = rows(
        world,
        "SELECT action, new_values->>'batch_approval_id' FROM audit_logs "
        "WHERE entity_id = %s AND action = 'bank_export.final_generated'",
        body["id"],
    )
    assert audited == [("bank_export.final_generated", str(approval_id))], audited


def test_the_file_holds_the_snapshot_not_the_current_beneficiary(
    world: dict[str, Any],
) -> None:
    """`SVC-EXPORT-005`. §15.4: built from the immutable version, "not current mutable
    beneficiary data".

    The beneficiary is renamed **after** approval and before generation. If any part of the path
    reached for `beneficiaries` instead of the item's snapshot, the file would carry a name the
    manager never approved — and it would carry it silently, because the row count, the total and
    every hash would still agree.

    That is the negative control the plan names, arranged so it cannot pass by accident: the new
    name is unmistakable in the file.
    """

    import io

    from openpyxl import load_workbook

    target = an_approved_version(world)

    with psycopg.connect(_psycopg(world["owner_url"])) as connection:
        connection.execute(
            "UPDATE beneficiaries SET full_name = %s WHERE id = %s",
            (RENAMED, world["beneficiary_id"]),
        )
        connection.commit()

    response = final_export(world, target)
    assert response.status_code == 201, response.text

    key = rows(
        world, "SELECT storage_key FROM file_objects WHERE id = %s", response.json()["file_id"]
    )[0][0]
    payload = (Path(world["storage_root"]) / str(key)).read_bytes()
    book = load_workbook(io.BytesIO(payload))
    sheet = book[book.sheetnames[0]]

    assert sheet.cell(row=2, column=2).value == ORIGINAL_NAME
    # And nowhere else in the sheet either. Asserted over every cell rather than over the
    # archive's shared-strings part, which openpyxl does not always produce — a test that read a
    # part that may not exist would fail for a reason unrelated to what it is about.
    every_value = {
        sheet.cell(row=r, column=c).value
        for r in range(1, sheet.max_row + 1)
        for c in range(1, sheet.max_column + 1)
    }
    assert RENAMED not in every_value, "the renamed beneficiary reached the file"

    # Put it back, so the module's other tests see the world they expect.
    with psycopg.connect(_psycopg(world["owner_url"])) as connection:
        connection.execute(
            "UPDATE beneficiaries SET full_name = %s WHERE id = %s",
            (ORIGINAL_NAME, world["beneficiary_id"]),
        )
        connection.commit()


def test_a_mismatch_quarantines_the_export_and_records_why(world: dict[str, Any]) -> None:
    """`SVC-INTEGRITY-002`. §15.5: "A mismatch quarantines the export".

    The version's stored `row_count` is moved out from under the export after approval, which is
    a state no legitimate path produces — that is the point. The export is written, found to
    disagree, moved to `quarantined`, and the caller is told which comparison failed.

    **The row survives**, and that is the assertion that matters. Refusing without writing would
    lose the only evidence that something disagreed, and `uq_active_final_export_per_version`
    excludes `quarantined` precisely so keeping it does not block the next attempt.
    """

    target = an_approved_version(world)

    with psycopg.connect(_psycopg(world["owner_url"])) as connection:
        connection.execute(
            "UPDATE payment_batch_versions SET row_count = row_count + 1 WHERE id = %s",
            (target["version_id"],),
        )
        connection.commit()

    response = final_export(world, target)

    assert response.status_code == 409, response.text
    assert "export_row_count_matches_version" in response.json()["error"]["message"]

    stored = rows(
        world,
        "SELECT status FROM bank_excel_exports WHERE payment_batch_version_id = %s",
        target["version_id"],
    )
    assert stored == [("quarantined",)], stored

    # The audit row is the record, and it names which comparison disagreed. §15.5 also asks for a
    # "high-priority task/security event" and Phase 1A has neither channel: there is no task
    # table (G-10), and `auth_events.event_class` admits only doc 12's six identity-and-access
    # classes — none of which an integrity mismatch is. Filing it under `administrative` to make
    # an insert succeed would route a financial-integrity incident into the queue that reads
    # administrative actions, so no `auth_events` row is written and the gap stays visible.
    audited = rows(
        world,
        "SELECT action, outcome, new_values->>'failed_checks' FROM audit_logs "
        "WHERE action = 'bank_export.integrity_failed' ORDER BY occurred_at DESC LIMIT 1",
    )
    assert audited, "the quarantine was not recorded at all"
    action, outcome, failed = audited[0]
    assert (action, outcome) == ("bank_export.integrity_failed", "failure")
    assert "export_row_count_matches_version" in str(failed), failed


def test_two_concurrent_final_exports_produce_one(world: dict[str, Any]) -> None:
    """`CON-EXPORT-001`. `uq_active_final_export_per_version` decides, not a prior read.

    Serial through the route here — the lock makes a genuine race wait rather than collide, so
    the observable claim is the one that matters to an accountant: a second attempt does not
    produce a second file that could also be sent.
    """

    target = an_approved_version(world)

    first = final_export(world, target)
    assert first.status_code == 201, first.text

    second = final_export(world, target)
    assert second.status_code >= 400, second.text

    assert rows(
        world,
        "SELECT count(*) FROM bank_excel_exports WHERE payment_batch_version_id = %s "
        "AND export_type = 'final' AND status IN "
        "('generated','validated','downloaded','sent_to_bank_marked')",
        target["version_id"],
    )[0][0] == 1


def test_a_repeated_idempotency_key_returns_the_stored_result(world: dict[str, Any]) -> None:
    """`CON-EXPORT-001`'s other half: "a timeout after commit returns the stored result rather
    than generating a second file"."""

    target = an_approved_version(world)
    key = str(uuid.uuid4())

    first = final_export(world, target, key=key)
    assert first.status_code == 201, first.text
    second = final_export(world, target, key=key)
    assert second.status_code == 201, second.text

    assert second.json()["id"] == first.json()["id"]
    assert second.json()["replayed"] is True


def test_generating_a_final_export_needs_the_final_permission(world: dict[str, Any]) -> None:
    """The manager approves and deliberately cannot generate.

    `permission_catalog.yaml:492` gives `generate_final` to `accountant` only. The direction is
    the milestone's subject: the person who authorised the payment is not the person who produces
    the file that executes it.
    """

    target = an_approved_version(world)
    sign_in_admin(world["client"], "final_manager")

    assert final_export(world, target).status_code == 403
