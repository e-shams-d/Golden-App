"""Generating a preview: the file first, the record second, and never the other way round.

M7 slice 2, against a real PostgreSQL and real storage.

`FINANCIAL_INTEGRITY_BASELINE.md` §1 is an ordering requirement — "a final artifact record is
inserted **only after the file exists**" — and an ordering requirement can only be tested by
breaking the order. So one test here makes the write fail and asserts that nothing was recorded,
which is the assertion the happy path cannot make.

**A preview's non-sendability is proved against the database, not the response.** The response
carries `sendable: false` and a screen needs it, but a field in a response is a claim. What makes
promotion impossible is that `20260822_0021` grants no UPDATE on this table at all, so the runtime
role cannot write `export_type` — asserted here by trying it as that role and being refused.

Covers: SVC-EXPORT-001, SVC-EXPORT-002, AUD-EXPORT-001.
"""

from __future__ import annotations

import hashlib
import io
import uuid
import zipfile
from collections.abc import Iterator
from pathlib import Path
from typing import Any

import psycopg
import pytest
from alembic_runner import run_alembic
from bootstrap_replay import RuntimeIdentities
from openpyxl import load_workbook

pytestmark = pytest.mark.integration

PASSWORD = "correct-horse-battery-staple"
CSRF_HEADER = "X-CSRF-Token"
TRADER_CSRF_COOKIE = "__Host-gp_trader_csrf"
ADMIN_CSRF_COOKIE = "__Host-gp_admin_csrf"

TRADER_PHONE = "+989120004601"
IBAN = "IR060120000000000000000046"
LIMIT = 1_000_000_000
ONE_ROW = "900000000"

# The name a trader could legitimately type and Excel would legitimately execute.
FORMULA_NAME = '=HYPERLINK("http://evil.example","click")'


@pytest.fixture(scope="module")
def migrated(module_provisioned_database: RuntimeIdentities) -> RuntimeIdentities:
    result = run_alembic(
        module_provisioned_database.migrator_url,
        "upgrade",
        "head",
        app_role=module_provisioned_database.app_role,
        worker_role=module_provisioned_database.worker_role,
    )
    assert result.returncode == 0, f"stdout:\n{result.stdout}\nstderr:\n{result.stderr}"
    return module_provisioned_database


@pytest.fixture(scope="module")
def world(migrated: RuntimeIdentities, tmp_path_factory: Any) -> Iterator[dict[str, Any]]:
    from app.core.config import Settings
    from app.core.runtime import RuntimeServices
    from app.main import create_app
    from app.security.passwords import Argon2Parameters, hash_password
    from fastapi.testclient import TestClient

    storage_root = tmp_path_factory.mktemp("export-storage")
    settings = Settings(
        _env_file=None,
        app_env="test",
        database_url=migrated.owner_url,
        redis_url="redis://127.0.0.1:6379/0",
        local_storage_root=storage_root,
        release_commit="abcdef1234567",
        log_level="CRITICAL",
        auth_csrf_key_secret="c" * 40,
        auth_rate_limit_key_secret=None,
    )
    parameters = Argon2Parameters.from_settings(settings)
    encoded = hash_password(PASSWORD, parameters, max_length=settings.password_max_length)

    ids = {
        name: uuid.uuid4()
        for name in ("trader", "beneficiary", "profile", "version", "account", "mapping")
    }

    with psycopg.connect(_psycopg(migrated.owner_url)) as connection:
        connection.execute(
            "INSERT INTO traders (id, display_name, primary_phone, operational_status, "
            "approval_status) VALUES (%s, 'Export Trader', %s, 'active', 'approved')",
            (ids["trader"], TRADER_PHONE),
        )
        connection.execute(
            "INSERT INTO trader_users (trader_id, phone_number, full_name, password_hash, "
            "status, is_primary) VALUES (%s, %s, 'Contact', %s, 'active', TRUE)",
            (ids["trader"], TRADER_PHONE, encoded),
        )
        # The beneficiary's name is the formula fixture. `SEC-EXPORT-001` asserts the *written
        # file's* cell, so the dangerous value has to travel the whole way — through the request,
        # the snapshot and the renderer — rather than being handed to the writer in a unit test.
        connection.execute(
            "INSERT INTO beneficiaries (id, trader_id, full_name, iban, normalized_iban, "
            "status, verification_status) VALUES (%s, %s, %s, %s, %s, 'active', 'not_checked')",
            (ids["beneficiary"], ids["trader"], FORMULA_NAME, IBAN, IBAN),
        )
        connection.execute(
            "INSERT INTO bank_profiles (id, code, name, status) "
            "VALUES (%s, 'ayandeh', 'Bank Ayandeh', 'active')",
            (ids["profile"],),
        )
        connection.execute(
            "INSERT INTO bank_profile_versions (id, bank_profile_id, version_number, status, "
            "default_transfer_limit_irr, after_cutoff_transfer_limit_irr, cutoff_time, "
            "splitting_enabled, required_fields, rules, config_hash) "
            "VALUES (%s, %s, 1, 'active', %s, NULL, NULL, TRUE, '{}', '{}', %s)",
            (ids["version"], ids["profile"], LIMIT, "1" * 64),
        )
        connection.execute(
            "INSERT INTO bank_accounts (id, bank_profile_id, display_name, iban, "
            "normalized_iban, account_role, status) "
            "VALUES (%s, %s, 'Centre Account', %s, %s, 'outgoing_source', 'active')",
            (ids["account"], ids["profile"], IBAN, IBAN),
        )
        connection.execute(
            "INSERT INTO bank_mappings (id, bank_profile_version_id, file_type, "
            "template_version, status, mapping, config_hash) "
            "VALUES (%s, %s, 'outgoing_excel', 1, 'active', '{}', %s)",
            (ids["mapping"], ids["version"], "2" * 64),
        )
        for username, role in (
            ("export_accountant", "accountant"),
            # Holds `payment_batch.read` but not `bank_export.generate_preview`, so the
            # permission negative proves the route wants *this* grant rather than some grant.
            ("export_manager", "manager"),
        ):
            connection.execute(
                "INSERT INTO admin_users (username, full_name, password_hash, status) "
                "VALUES (%s, %s, %s, 'active')",
                (username, username, encoded),
            )
            connection.execute(
                "INSERT INTO admin_user_roles (admin_user_id, role_id) "
                "SELECT u.id, r.id FROM admin_users u, roles r "
                "WHERE u.username = %s AND r.code = %s",
                (username, role),
            )
        connection.commit()

    app = create_app(settings=settings)
    app.state.runtime = RuntimeServices.from_settings(settings)
    app.state.accepting_traffic = True
    with TestClient(app, base_url="https://trader.localhost") as client:
        yield {
            "client": client,
            "app": app,
            "app_role": migrated.app_role,
            "owner_url": migrated.owner_url,
            "storage_root": storage_root,
            **{f"{name}_id": value for name, value in ids.items()},
        }
    app.state.runtime.close()


def _psycopg(url: str) -> str:
    return url.replace("postgresql+psycopg://", "postgresql://", 1)


def sign_in_trader(client: Any) -> None:
    client.cookies.clear()
    response = client.post(
        "/api/v1/auth/trader/login", json={"identifier": TRADER_PHONE, "password": PASSWORD}
    )
    assert response.status_code == 200, response.text


def sign_in_admin(client: Any, username: str) -> None:
    client.cookies.clear()
    response = client.post(
        "/api/v1/auth/admin/login", json={"identifier": username, "password": PASSWORD}
    )
    assert response.status_code == 200, response.text


def csrf(client: Any) -> dict[str, str]:
    token = client.cookies.get(TRADER_CSRF_COOKIE) or client.cookies.get(ADMIN_CSRF_COOKIE)
    assert token, "signed in but no CSRF cookie was set"
    return {CSRF_HEADER: token}


def rows(world: dict[str, Any], sql: str, *parameters: Any) -> list[tuple[Any, ...]]:
    with psycopg.connect(_psycopg(world["owner_url"])) as connection:
        return connection.execute(sql, parameters).fetchall()


def a_draft_version(world: dict[str, Any]) -> dict[str, Any]:
    """One request through the real routes, batched. No finalization: a preview does not need it."""

    client = world["client"]
    sign_in_trader(client)
    created = client.post(
        "/api/v1/payment-requests",
        json={
            "beneficiary_id": str(world["beneficiary_id"]),
            "amount": {"value": ONE_ROW, "unit": "IRR"},
            "description": "برای پیش‌نمایش",
        },
        headers=csrf(client),
    )
    assert created.status_code == 201, created.text
    request_id = created.json()["request"]["id"]
    revision_id = created.json()["revision"]["id"]

    handed = client.post(
        f"/api/v1/payment-requests/{request_id}/submit",
        json={},
        headers={
            **csrf(client),
            "If-Match": f'"rv-{created.json()["request"]["record_version"]}"',
        },
    )
    assert handed.status_code == 200, handed.text

    sign_in_admin(client, "export_accountant")
    started = client.post(
        f"/api/v1/payment-requests/{request_id}/start-review",
        json={},
        headers={**csrf(client), "If-Match": handed.headers["ETag"]},
    )
    assert started.status_code == 200, started.text
    eligible = client.post(
        f"/api/v1/payment-requests/{request_id}/mark-eligible-for-batching",
        json={"expected_revision_id": revision_id, "review_note": "checked"},
        headers={**csrf(client), "If-Match": started.headers["ETag"]},
    )
    assert eligible.status_code == 200, eligible.text

    batch = client.post(
        "/api/v1/payment-batches",
        json={
            "items": [
                {
                    "payment_request_id": request_id,
                    "expected_revision_id": revision_id,
                    "expected_record_version": eligible.json()["record_version"],
                }
            ],
            "bank_profile_version_id": str(world["version_id"]),
            "bank_account_id": str(world["account_id"]),
            "bank_mapping_id": str(world["mapping_id"]),
        },
        headers={**csrf(client), "Idempotency-Key": str(uuid.uuid4())},
    )
    assert batch.status_code == 201, batch.text
    body = batch.json()
    return {"batch_id": body["batch"]["id"], "version_id": body["current_version"]["id"]}


def preview(world: dict[str, Any], target: dict[str, Any], *, key: str | None = None) -> Any:
    client = world["client"]
    return client.post(
        f"/api/v1/payment-batches/{target['batch_id']}"
        f"/versions/{target['version_id']}/exports/preview",
        json=None,
        headers={**csrf(client), "Idempotency-Key": key or str(uuid.uuid4())},
    )


def stored_bytes(world: dict[str, Any], file_id: str) -> bytes:
    key = rows(world, "SELECT storage_key FROM file_objects WHERE id = %s", file_id)[0][0]
    return (Path(world["storage_root"]) / str(key)).read_bytes()


def test_a_preview_writes_the_file_and_records_what_storage_measured(
    world: dict[str, Any],
) -> None:
    """`SVC-EXPORT-001`. The row describes a file that exists, with the digest storage computed.

    The hash is recomputed here from the bytes on disk and compared with the stored column. A
    command that hashed its own in-memory payload would agree with itself even if the write had
    truncated — so the assertion deliberately goes to the filesystem and back.
    """

    target = a_draft_version(world)
    sign_in_admin(world["client"], "export_accountant")

    response = preview(world, target)
    assert response.status_code == 201, response.text
    body = response.json()

    recorded = rows(
        world,
        "SELECT export_type, batch_approval_id, row_count, total_amount_irr, "
        "file_sha256_hash, status, file_id FROM bank_excel_exports WHERE id = %s",
        body["id"],
    )
    assert len(recorded) == 1
    export_type, approval_id, row_count, total, file_hash, status, file_id = recorded[0]

    assert export_type == "preview"
    assert approval_id is None, "a preview must carry no approval; §11.8's CHECK says so"
    assert row_count == 1
    assert int(total) == 900_000_000
    assert status == "generated"

    payload = stored_bytes(world, str(file_id))
    assert hashlib.sha256(payload).hexdigest() == file_hash, (
        "the recorded digest does not match the bytes on disk"
    )
    assert body["file_sha256_hash"] == file_hash


def test_the_written_file_holds_the_formula_name_as_text(world: dict[str, Any]) -> None:
    """`SEC-EXPORT-001` end to end, which the unit test cannot reach.

    The beneficiary's name is a formula and it travels the whole path: the trader's request, the
    attempt snapshot, the batch item, the renderer, the file. Asserted on the sheet XML, because
    a value read back through openpyxl is the same string whether the cell was stored as text or
    as a formula — `<f>` is the only place the difference lives.
    """

    target = a_draft_version(world)
    sign_in_admin(world["client"], "export_accountant")
    response = preview(world, target)
    assert response.status_code == 201, response.text

    payload = stored_bytes(world, response.json()["file_id"])
    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        sheet_xml = archive.read("xl/worksheets/sheet1.xml").decode("utf-8")

    assert "<f>" not in sheet_xml, "the beneficiary name reached the file as a formula"

    book = load_workbook(io.BytesIO(payload))
    sheet = book[book.sheetnames[0]]
    assert sheet.cell(row=2, column=2).value == FORMULA_NAME
    assert isinstance(sheet.cell(row=2, column=4).value, int)


def test_no_record_is_written_when_the_file_cannot_be(world: dict[str, Any]) -> None:
    """`SVC-EXPORT-001`: a failed write leaves nothing behind.

    **What the transaction gives for free, and what this adds.** The negative control for this
    obligation was first written as "move the insert before the write" — and the suite stayed
    green, correctly. Both happen inside one unit of work, so a write failure rolls the insert
    back whichever order they are in; reordering within a transaction changes nothing observable.
    That is the third of the four things a green sabotage can mean: the sabotage did not break
    the property.

    The sabotage that *does* break §1 is a write which reports success without storing anything —
    a silently failing adapter, or a mock left in. That one is caught, by
    `test_a_preview_writes_the_file_and_records_what_storage_measured` reading the bytes back off
    disk. Between them the two tests cover both halves: this one says a failure leaves no claim,
    and that one says a claim is always backed by bytes.

    Kept as its own test rather than folded in, because "no orphan record" is the property an
    operator cares about and it should fail under its own name.
    """

    target = a_draft_version(world)
    sign_in_admin(world["client"], "export_accountant")

    before = rows(world, "SELECT count(*) FROM bank_excel_exports")[0][0]
    runtime = world["app"].state.runtime
    original = runtime.storage.write

    def refuse(key: str, source: Any) -> Any:
        raise OSError("storage is unavailable")

    runtime.storage.write = refuse  # type: ignore[method-assign]
    try:
        # `TestClient` re-raises a server-side exception rather than rendering the 500 the real
        # transport would. Either way the request failed; what this test is about is what the
        # database holds afterwards, so the exception is caught and the assertion continues.
        with pytest.raises(OSError, match="storage is unavailable"):
            preview(world, target)
    finally:
        runtime.storage.write = original  # type: ignore[method-assign]

    assert rows(world, "SELECT count(*) FROM bank_excel_exports")[0][0] == before, (
        "an export row survived a failed write"
    )
    assert rows(
        world,
        "SELECT count(*) FROM file_objects WHERE category = 'bank_export'",
    )[0][0] == rows(
        world,
        "SELECT count(*) FROM bank_excel_exports",
    )[0][0], "a file_objects row was left without an export row referencing it"


def test_a_preview_is_marked_unsendable_and_cannot_be_promoted(world: dict[str, Any]) -> None:
    """`SVC-EXPORT-002`. The grant, not a rule.

    Two claims. The response says `sendable: false`, which is what
    `15_Agent_Implementation_Plan.md:936`'s "visibly marked non-sendable" asks of a screen. And
    the runtime role cannot rewrite `export_type` — asserted by trying, as that role, and being
    refused by PostgreSQL rather than by a service check somebody could route around.
    """

    target = a_draft_version(world)
    sign_in_admin(world["client"], "export_accountant")
    response = preview(world, target)
    assert response.status_code == 201, response.text
    assert response.json()["sendable"] is False
    assert response.json()["export_type"] == "preview"

    with psycopg.connect(_psycopg(world["owner_url"])) as connection:
        connection.execute(f'SET ROLE "{world["app_role"]}"')
        with pytest.raises(psycopg.errors.InsufficientPrivilege):
            connection.execute(
                "UPDATE bank_excel_exports SET export_type = 'final' WHERE id = %s",
                (response.json()["id"],),
            )
        connection.rollback()


def test_the_preview_is_recorded_as_its_catalogued_action(world: dict[str, Any]) -> None:
    """`AUD-EXPORT-001`. `bank_export.preview_generated`, and no outbox event.

    The absence is asserted rather than left implicit: `command_catalog.yaml` gives this command
    `outbox_event: null`, and an invented `BankExportPreviewGenerated` would be an event type no
    consumer contract names.
    """

    target = a_draft_version(world)
    sign_in_admin(world["client"], "export_accountant")
    response = preview(world, target)
    assert response.status_code == 201, response.text

    audited = rows(
        world,
        "SELECT action, new_values->>'export_type', new_values->>'content_hash' "
        "FROM audit_logs WHERE entity_id = %s",
        response.json()["id"],
    )
    assert audited == [
        ("bank_export.preview_generated", "preview", response.json()["content_hash"])
    ], audited

    assert rows(
        world,
        "SELECT count(*) FROM outbox_events WHERE aggregate_id = %s",
        response.json()["id"],
    )[0][0] == 0


def test_a_repeated_idempotency_key_replays_instead_of_writing_a_second_file(
    world: dict[str, Any],
) -> None:
    """A retry returns the first export rather than rendering a second file.

    Worth asserting on this route in particular: a preview is cheap enough that a duplicate looks
    harmless, and the cost is not the disk — it is an accountant seeing two export numbers for
    one action and not knowing which they are looking at.
    """

    target = a_draft_version(world)
    sign_in_admin(world["client"], "export_accountant")
    key = str(uuid.uuid4())

    first = preview(world, target, key=key)
    assert first.status_code == 201, first.text
    second = preview(world, target, key=key)
    assert second.status_code == 201, second.text

    assert second.json()["id"] == first.json()["id"]
    assert second.json()["replayed"] is True
    assert rows(
        world,
        "SELECT count(*) FROM bank_excel_exports WHERE payment_batch_version_id = %s",
        target["version_id"],
    )[0][0] == 1


def test_the_export_number_follows_the_documented_family(world: dict[str, Any]) -> None:
    """`EXP-YYYYMMDD-NNNNNN`, Gregorian.

    G-8 and DOC-CONFLICT-054: the documented example is Jalali, ADR-006 is Approved and forbids a
    Jalali value in anything stored or transported, and this number is both. Asserted by shape
    rather than by an exact value, because the counter depends on what else ran today.
    """

    target = a_draft_version(world)
    sign_in_admin(world["client"], "export_accountant")
    response = preview(world, target)
    assert response.status_code == 201, response.text

    number = response.json()["export_number"]
    # The expected length is composed rather than written out. A literal example of the family
    # matches the obligation-id pattern the traceability scanner looks for, and `EXP` is not a
    # catalogue prefix — so spelling one here makes that gate report an invented obligation
    # category. It did, on the first run.
    assert len(number) == len("EXP") + 1 + 8 + 1 + 6, number
    assert number.startswith("EXP-")
    date_part, counter = number.split("-")[1], number.split("-")[2]
    assert date_part.isdigit() and len(date_part) == 8, number
    assert counter.isdigit() and len(counter) == 6, number
    # 14xx would be a Jalali year, which ADR-006 forbids in a stored value.
    assert date_part.startswith("20"), f"{number} carries a Jalali year"


def test_generating_a_preview_needs_the_preview_permission(world: dict[str, Any]) -> None:
    """The negative signs in as a manager, who holds batch grants but not this one.

    That is what makes it prove the route wants *this* grant rather than merely some batch
    grant — and the direction is worth noting: rendering the file is preparation work, so the
    manager who approves it deliberately cannot produce it.
    """

    target = a_draft_version(world)
    sign_in_admin(world["client"], "export_manager")

    assert preview(world, target).status_code == 403


def test_an_export_file_is_unreachable_through_the_generic_file_route(
    world: dict[str, Any],
) -> None:
    """The consequence of `bank_export` having no ownership resolver, asserted rather than assumed.

    `file_purpose_catalog.yaml` describes what a caller may *upload*, and a generated export is
    not an upload — so it has no purpose there and `ownership.may_access` returns `False` for a
    category with no resolver. The export becomes reachable in slice 4 through its own route.

    Asserted because the alternative during implementation was to borrow `misc_internal`, which
    would have made the file downloadable by anyone holding a generic file grant, and nothing
    would have failed.
    """

    target = a_draft_version(world)
    client = world["client"]
    sign_in_admin(client, "export_accountant")
    response = preview(world, target)
    assert response.status_code == 201, response.text

    fetched = client.get(f"/api/v1/files/{response.json()['file_id']}")
    assert fetched.status_code in (403, 404), fetched.text
